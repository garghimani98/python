from openpyxl import load_workbook
import pandas as pd
import openpyxl as xl
import numpy as np



filename1="C:\\Users\\Users\\Documents\\File_name_with_extension"
wb1=xl.load_workbook(filename1)

ws1=wb1.active

filename2="C:\\Users\\Users\\Documents\\File_name_with_extension"
wb2=xl.load_workbook(filename2)


ws2=wb2.active



#for unmatching rows
filename4="C:\\Users\\Users\\Documents\\File_name_with_extension"
wb4=xl.load_workbook(filename4)
wb4.remove(wb4.active)
ws4 = wb4.create_sheet("sheet1")
#ws4=wb4.active

#number of rows and columns
m_r1=ws1.max_row
m_c1=ws1.max_column
m_r2=ws2.max_row
m_c2=ws2.max_column
     
#for appending columns
row_data=[]
for j in range(1,m_c1+1,1):
       cell=ws1.cell(row=1, column=j).value
       row_data.append(cell)
#ws3.append(row_data) 
ws4.append(row_data)     
x=True
    
if(m_c1==m_c2):
    
    for m in range(2,m_r1+1,1):
      y=False  
      for n in range(2,m_r2+1,1):
          for i in range(1,m_c1+1,1) :
              h=ws1.cell(row=m,column=i).value
              
              g=ws2.cell(row=n,column=i).value
              x=h==g
              if x==False:
                break
        
          if x==True:
              y=True
              #row_data1=[]
              #for j in range(1,m_c1+1):
                #cell=ws1.cell(row=m, column=j).value
                #row_data1.append(cell)
              #ws3.append(row_data1)
              
              #print(row_data1)
              break
          else:
            continue
      if y==False:
          row_data2=[]
          for j in range(1,m_c1+1):
              cell=ws1.cell(row=m, column=j).value
              row_data2.append(cell)
          ws4.append(row_data2)
wb4.save(filename4) 
#wb3.save(filename3) 
#####################################              
df1=pd.read_excel(r"C:\\Users\\Users\\Documents\\File_name_with_extension")


   
#to get number of plants in  excel 
df1 = df1.sort_values(by ='Plant')
index = df1.index
n_r = len(index)


##to add values to list
plant=[]
for row in df1['Plant']:
    #print(row)
    if(len(plant)!=0):
        if(plant[-1]==row):
           continue
        else:
           plant.append(row) 
    else:      
      plant.append(row)
 
## to apply filters on columns   
for i in range(0,len(plant)):
     #res.append(row_data)
     res= df1[df1['Plant'] == plant[i]] 
     print(res)
     workbook_name="C:\\Users\\Users\\Documents\\"+plant[i]+".xlsx" ##PATH FOR STORING XLSX FILE
     wb = xl.Workbook()
     wb.remove(wb.active)
     ws = wb.create_sheet("sheet1")
     res.to_excel(workbook_name,index = False,header=True)
     print(ws)
       
        
             
#wb3.delete_cols(1, 4)
#wb3.delete_rows(1, 10)
#wb3.close()







