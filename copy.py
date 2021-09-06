import pandas as pd
import openpyxl as xl

filename ="C:\\Users\\Users\\Documents\\File_name_with_extension"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.active

filename1="C:\\Users\\Users\\Documents\\File_name_with_extension"
wb2=xl.load_workbook(filename1)
ws2=wb2.active

m_r=ws1.max_row
print(m_r)
m_c = ws1.max_column

for i in range(1,m_r+1):
    for j in range(1,m_c+1):
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
        
         # writing the read value to destination excel file 
        ws2.cell(row = i, column = j).value = c.value 
wb2.save(str(filename1))

######################################
#print(dropped_rows)
#added_rows = set(new.index) - set(old.index)

#to append column labels 
row_data=[]
for col in old.columns:
       row_data.append(col)
ws3.append(row_data) 

for row in dropped_rows:
    print(row)
    #ws3.append(row)
    
#To save the this excel sheet    
#wb3.save(filename3) 