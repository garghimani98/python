import openpyxl as xl

workbook = xl.load_workbook(filename="C:\\Users\\Users\\Documents\\File_name_with_extension")

title_column_name = "title"

# Get the active worksheet
ws = workbook.active

# The String we'll search for. You could prompt the user to provide
# this using python2's raw_input, oder python3's input function.
searchstring = "feb"
m_r=ws.max_row
# ws.rows[1:] means we'll skip the first row (the header row).
for i in range(1,m_r+1):
    # row[1] is the title column. string.find(str) returns -1
    # if the value was not found, or the index in the string if
    # the value was found.
    if ws[i][1].value.find(searchstring) != -1:
        print("Found a matching row! month={0}, note={1}".format(ws[i][1].value, ws[i][2].value))